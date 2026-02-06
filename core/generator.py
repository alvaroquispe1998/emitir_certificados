from __future__ import annotations

import re
import tempfile
from io import BytesIO
from pathlib import Path, PurePosixPath
import posixpath
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import yaml
from openpyxl import load_workbook

from .io import (
    read_csv,
    normalize_str,
    extract_course_code,
    validate_required_columns,
    parse_grade,
    parse_year,
)
from .model_indexer import build_model_index
from .electives import balance_electives, parse_period
from .logger import LogCollector
from .zipout import zip_directory


INVALID_FILENAME_CHARS = r"\\/:*?\"<>|"
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
REL_TYPE_DRAWING = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"

ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_REL)
COURSE_TITLE_STOPWORDS = {
    "a",
    "al",
    "con",
    "de",
    "del",
    "el",
    "en",
    "la",
    "las",
    "los",
    "o",
    "para",
    "por",
    "sin",
    "u",
    "y",
    "e",
}


def _load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _is_missing(value: object) -> bool:
    return value is None or str(value).strip() == ""


def _sanitize_filename(text: str) -> str:
    cleaned = re.sub("[{0}]".format(re.escape(INVALID_FILENAME_CHARS)), "", text)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _is_valid_dni(dni: str) -> bool:
    return bool(dni) and dni.isdigit()


def _normalize_col_name(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).strip().upper())


def _normalize_spaces(text: str) -> str:
    return " ".join(str(text or "").strip().split())


def _split_words(text: str) -> list[str]:
    return [w for w in _normalize_spaces(text).split(" ") if w]


def _is_roman_numeral(token: str) -> bool:
    token = token.upper()
    if not token:
        return False
    return bool(re.fullmatch(r"M{0,4}(CM|CD|D?C{0,3})(XC|XL|L?X{0,3})(IX|IV|V?I{0,3})", token))


def _format_word(token: str) -> str:
    if not token:
        return token
    start = 0
    end = len(token) - 1
    while start <= end and not token[start].isalnum():
        start += 1
    while end >= start and not token[end].isalnum():
        end -= 1
    if start > end:
        return token

    leading = token[:start]
    core = token[start : end + 1]
    trailing = token[end + 1 :]

    if core.isdigit():
        return leading + core + trailing
    if _is_roman_numeral(core):
        return leading + core.upper() + trailing
    if core.lower() in COURSE_TITLE_STOPWORDS:
        return leading + core.lower() + trailing
    if core.isupper() and len(core) <= 3:
        return leading + core.upper() + trailing
    return leading + core[0].upper() + core[1:].lower() + trailing


def _format_course_name(name: str) -> str:
    name = _normalize_spaces(name)
    if not name:
        return ""
    parts = re.split(r"([\s\-\/])", name)
    formatted = []
    for part in parts:
        if part in {" ", "-", "/"}:
            formatted.append(part)
            continue
        formatted.append(_format_word(part))
    return "".join(formatted)


def _format_period_value(value: object) -> str:
    period = parse_period(value)
    if period is None:
        return ""
    return str(period)


def _parse_resolution_flag(value: object) -> tuple[bool, str | None]:
    text = normalize_str(value)
    if not text:
        return True, None
    if text in {"SI", "SÍ", "S"}:
        return True, None
    if text in {"NO", "N"}:
        return False, None
    return True, "Valor inválido en columna RESOLUCION: '{0}' (use SI o NO)".format(value)


def _format_correlativo(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.isdigit():
        return text.zfill(3)
    return text


def _elective_correlativo_from_code(code: str) -> str | None:
    code = normalize_str(code)
    if not code:
        return None
    match = re.search(r"AA(\d+)$", code)
    if not match:
        return None
    try:
        num = int(match.group(1))
    except ValueError:
        return None
    if 1 <= num <= 6:
        return "{0:03d}".format(48 + num)
    return None


def _resolve_correlativo(rec: dict) -> str:
    if normalize_str(rec.get("TIPO_CURSO")) == "ELECTIVO":
        derived = _elective_correlativo_from_code(rec.get("CODIGO_CURSO") or "")
        if derived:
            return derived
    return _format_correlativo(rec.get("Z"))


def _format_simple_person_name(name: str) -> tuple[str | None, str | None]:
    name = _normalize_spaces(name)
    if not name:
        return None, "Nombre vacío"
    return name.upper(), None


def _load_name_overrides(config: dict, config_path: str) -> tuple[dict[str, int], Path | None]:
    overrides_cfg = config.get("name_overrides") or {}
    path = overrides_cfg.get("path") or "name_overrides.csv"
    overrides_path = Path(path)
    if not overrides_path.is_absolute():
        overrides_path = Path(config_path).parent / overrides_path
    if not overrides_path.exists():
        return {}, overrides_path

    df = read_csv(overrides_path)
    if df.empty:
        return {}, overrides_path

    dni_col = _resolve_column(df, overrides_cfg.get("dni_column"), ["DNI"])
    comma_col = _resolve_column(
        df, overrides_cfg.get("comma_after_column"), ["COMMA_AFTER", "COMA_DESPUES"]
    )
    if not dni_col or not comma_col:
        return {}, overrides_path

    mapping: dict[str, int] = {}
    for rec in df.to_dict("records"):
        dni = str(rec.get(dni_col) or "").strip()
        if not dni:
            continue
        try:
            comma_after = int(str(rec.get(comma_col)).strip())
        except Exception:
            continue
        mapping[dni] = comma_after
    return mapping, overrides_path


def _write_name_overrides(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    import pandas as pd

    existing = None
    if path.exists():
        try:
            existing = read_csv(path)
        except Exception:
            existing = None

    df_new = pd.DataFrame(rows, columns=["DNI", "NOMBRE_COMPLETO", "COMMA_AFTER"])
    if existing is not None and not existing.empty:
        merged = pd.concat([existing, df_new], ignore_index=True)
        merged = merged.drop_duplicates(subset=["DNI"], keep="first")
        merged.to_csv(path, index=False)
        return
    df_new.to_csv(path, index=False)


def _format_person_name(
    name: str,
    dni: str,
    overrides: dict[str, int],
    ambiguous_out: list[dict],
) -> tuple[str | None, str | None]:
    name = _normalize_spaces(name)
    if not name:
        return None, "Nombre vacío"

    words = _split_words(name)
    upper_name = name.upper()
    has_de_la = "DE LA" in upper_name
    word_count = len(words)

    override_after = overrides.get(dni)
    if override_after:
        if 1 <= override_after < word_count:
            left = " ".join(words[:override_after])
            right = " ".join(words[override_after:])
            return "{0}, {1}".format(left, right), None
        return None, "COMMA_AFTER inválido para DNI: {0}".format(dni)

    if has_de_la or word_count not in (3, 4):
        ambiguous_out.append(
            {"DNI": dni, "NOMBRE_COMPLETO": name, "COMMA_AFTER": ""}
        )
        return None, "Nombre requiere coma manual: {0}".format(name)

    split_after = 1 if word_count == 3 else 2
    left = " ".join(words[:split_after])
    right = " ".join(words[split_after:])
    return "{0}, {1}".format(left, right), None


def _resolve_column(df, preferred: str | None, aliases: list[str]) -> str | None:
    if preferred:
        target = _normalize_col_name(preferred)
        for col in df.columns:
            if _normalize_col_name(col) == target:
                return col
    for alias in aliases:
        target = _normalize_col_name(alias)
        for col in df.columns:
            if _normalize_col_name(col) == target:
                return col
    return None


def _load_course_metadata(config: dict, config_path: str):
    meta_cfg = config.get("course_metadata") or {}
    path = meta_cfg.get("path")
    if not path:
        return {}

    meta_path = Path(path)
    if not meta_path.is_absolute():
        meta_path = Path(config_path).parent / meta_path

    df = read_csv(meta_path)
    if df.empty:
        return {}

    code_col = _resolve_column(df, meta_cfg.get("code_column"), ["CODIGO_CURSO", "CODIGO CURSO", "CODIGO"])
    res_col = _resolve_column(df, meta_cfg.get("resolucion_column"), ["RESOLUCION", "RESOLUCIÓN"])
    z_col = _resolve_column(df, meta_cfg.get("z_column"), ["Z"])

    missing = []
    if not code_col:
        missing.append("CODIGO_CURSO")
    if not res_col:
        missing.append("RESOLUCION")
    if not z_col:
        missing.append("Z")
    if missing:
        raise ValueError(
            "Faltan columnas en course_metadata: {0}".format(", ".join(missing))
        )

    mapping: dict[str, dict[str, str]] = {}
    for rec in df.to_dict("records"):
        code = extract_course_code(rec.get(code_col))
        if not code:
            continue
        mapping[code] = {
            "resolucion": str(rec.get(res_col) or "").strip(),
            "z": str(rec.get(z_col) or "").strip(),
        }

    return mapping


def _get_course_columns(sheet_cfg: dict) -> dict[str, str]:
    cols = sheet_cfg.get("course_columns") or {}

    def _norm_col(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip().upper()
        return text if text and text != "NONE" else ""

    return {
        "code_col": _norm_col(cols.get("code_col", "B")),
        "curso_col": _norm_col(cols.get("curso_col", "")),
        "nota_col": _norm_col(cols.get("nota_col", "E")),
        "nota_texto_col": _norm_col(cols.get("nota_texto_col", "F")),
        "creditos_col": _norm_col(cols.get("creditos_col", "")),
        "acta_col": _norm_col(cols.get("acta_col", "")),
        "resolucion_col": _norm_col(cols.get("resolucion_col", "")),
        "x_col": _norm_col(cols.get("x_col", "")),
        "y_col": _norm_col(cols.get("y_col", "")),
        "z_col": _norm_col(cols.get("z_col", "")),
    }


def _select_sheet(year_int: int | None, cara_sheet: str, sello_sheet: str) -> str | None:
    if year_int is None:
        return None
    if 1 <= year_int <= 5:
        return cara_sheet
    if 6 <= year_int <= 10:
        return sello_sheet
    return None


def _write_identity(ws, identity, header_cells):
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    if not header_cells:
        raise ValueError("No se configuraron celdas de cabecera")

    required = {
        "dni": "DNI",
        "nombre": "NOMBRE",
        "facultad": "FACULTAD",
    }

    positions = {}
    missing = []

    for key, identity_key in required.items():
        coord = header_cells.get(key)
        if not coord:
            missing.append(key.upper())
            continue
        try:
            col_letter, row = coordinate_from_string(coord)
            col = column_index_from_string(col_letter)
            positions[identity_key] = (row, col)
        except Exception:
            missing.append(key.upper())

    program_coord = header_cells.get("escuela") or header_cells.get("programa")
    if not program_coord:
        missing.append("ESCUELA")
    else:
        try:
            col_letter, row = coordinate_from_string(program_coord)
            col = column_index_from_string(col_letter)
            positions["PROGRAMA"] = (row, col)
        except Exception:
            missing.append("ESCUELA")

    date_coord = header_cells.get("fecha")
    if date_coord:
        try:
            col_letter, row = coordinate_from_string(date_coord)
            col = column_index_from_string(col_letter)
            positions["FECHA"] = (row, col)
        except Exception:
            missing.append("FECHA")

    if missing:
        raise ValueError("No se pudo ubicar campos de cabecera: {0}".format(", ".join(missing)))

    for key, (row, col) in positions.items():
        ws.cell(row=row, column=col).value = identity.get(key, "")


def _merged_anchor(merged_ranges, row: int, col: int) -> tuple[int, int]:
    if not merged_ranges:
        return row, col
    for cell_range in merged_ranges:
        if (
            cell_range.min_row <= row <= cell_range.max_row
            and cell_range.min_col <= col <= cell_range.max_col
        ):
            return cell_range.min_row, cell_range.min_col
    return row, col


def _safe_set_cell(ws, row: int, col: int | None, value, merged_ranges) -> None:
    if not col:
        return
    target_row, target_col = _merged_anchor(merged_ranges, row, col)
    ws.cell(row=target_row, column=target_col).value = value


def _grade_for_output(grade: int | None):
    if grade == 0:
        return None
    return grade


def _resolution_output_value(
    rec: dict, x_ref: str | None = None, y_ref: str | None = None, z_ref: str | None = None
) -> str:
    if rec.get("USA_RESOLUCION", True):
        return rec.get("RESOLUCION_TEXTO", "")
    if x_ref and y_ref and z_ref:
        return "=+CONCATENATE({0},{1},{2})".format(x_ref, y_ref, z_ref)
    return "{0}{1}{2}".format(
        rec.get("PERIODO_VAL", ""),
        "21",
        rec.get("CORRELATIVO", ""),
    )


def _clear_course_row(ws, row: int, cols: dict[str, int | None], merged_ranges) -> None:
    for key in (
        "code_col",
        "curso_col",
        "nota_col",
        "nota_texto_col",
        "creditos_col",
        "acta_col",
        "resolucion_col",
        "x_col",
        "y_col",
        "z_col",
    ):
        _safe_set_cell(ws, row, cols.get(key), None, merged_ranges)


def _worksheet_rel_path(sheet_path: str) -> str:
    prefix, name = sheet_path.rsplit("/", 1)
    return "{0}/_rels/{1}.rels".format(prefix, name)


def _resolve_part(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = PurePosixPath(base_part).parent
    return posixpath.normpath(str(PurePosixPath(base_dir, target)))


def _sheet_paths_from_workbook(template_entries: dict[str, bytes]) -> dict[str, str]:
    workbook = ET.fromstring(template_entries["xl/workbook.xml"])
    rels = ET.fromstring(template_entries["xl/_rels/workbook.xml.rels"])
    rel_targets: dict[str, str] = {}
    for rel in rels.findall("{%s}Relationship" % NS_PKG_REL):
        rel_targets[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")

    paths: dict[str, str] = {}
    for sheet in workbook.findall(".//{%s}sheet" % NS_MAIN):
        name = sheet.attrib.get("name", "")
        rid = sheet.attrib.get("{%s}id" % NS_REL, "")
        target = rel_targets.get(rid, "")
        if not name or not target:
            continue
        paths[name] = _resolve_part("xl/workbook.xml", target)
    return paths


def _apply_date_to_drawing_xml(xml_bytes: bytes, certificate_date: str) -> bytes:
    if not certificate_date.strip():
        return xml_bytes
    root = ET.fromstring(xml_bytes)
    shape_nodes = root.findall(".//{%s}sp" % NS_XDR)
    for shape in shape_nodes:
        texts = shape.findall(".//{%s}t" % NS_A)
        merged = "".join([(node.text or "") for node in texts]).strip().lower()
        if not merged:
            continue
        if re.search(r"\d{1,2}\s+de\s+.+\s+del\s+\d{4}", merged):
            texts[0].text = certificate_date
            for extra in texts[1:]:
                extra.text = ""
            return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    if shape_nodes:
        texts = shape_nodes[0].findall(".//{%s}t" % NS_A)
        if texts:
            texts[0].text = certificate_date
            for extra in texts[1:]:
                extra.text = ""
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _copy_template_drawings(
    template_path: Path, output_path: Path, certificate_date: str
) -> None:
    with ZipFile(template_path, "r") as z_tpl:
        template_entries = {name: z_tpl.read(name) for name in z_tpl.namelist()}
    with ZipFile(output_path, "r") as z_out:
        out_entries = {name: z_out.read(name) for name in z_out.namelist()}

    sheet_paths = _sheet_paths_from_workbook(template_entries)
    drawing_parts: set[str] = set()

    for _sheet_name, sheet_path in sheet_paths.items():
        tpl_sheet_xml = template_entries.get(sheet_path)
        out_sheet_xml = out_entries.get(sheet_path)
        if not tpl_sheet_xml or not out_sheet_xml:
            continue

        tpl_sheet = ET.fromstring(tpl_sheet_xml)
        out_sheet = ET.fromstring(out_sheet_xml)

        tpl_drawings = tpl_sheet.findall("{%s}drawing" % NS_MAIN)
        if not tpl_drawings:
            continue

        existing_rids = {
            node.attrib.get("{%s}id" % NS_REL, "")
            for node in out_sheet.findall("{%s}drawing" % NS_MAIN)
        }
        changed_sheet = False
        for tpl_drawing in tpl_drawings:
            rid = tpl_drawing.attrib.get("{%s}id" % NS_REL, "")
            if rid in existing_rids:
                continue
            out_sheet.append(ET.fromstring(ET.tostring(tpl_drawing, encoding="utf-8")))
            changed_sheet = True
        if changed_sheet:
            out_entries[sheet_path] = ET.tostring(out_sheet, encoding="utf-8", xml_declaration=True)

        rel_path = _worksheet_rel_path(sheet_path)
        tpl_rels_xml = template_entries.get(rel_path)
        if not tpl_rels_xml:
            continue

        tpl_rels = ET.fromstring(tpl_rels_xml)
        out_rels = ET.fromstring(out_entries[rel_path]) if rel_path in out_entries else ET.Element(
            "{%s}Relationships" % NS_PKG_REL
        )

        existing_rel_ids = {
            rel.attrib.get("Id", "") for rel in out_rels.findall("{%s}Relationship" % NS_PKG_REL)
        }
        rels_changed = False
        for rel in tpl_rels.findall("{%s}Relationship" % NS_PKG_REL):
            if rel.attrib.get("Type") != REL_TYPE_DRAWING:
                continue
            rel_id = rel.attrib.get("Id", "")
            target = rel.attrib.get("Target", "")
            if rel_id not in existing_rel_ids:
                out_rels.append(ET.fromstring(ET.tostring(rel, encoding="utf-8")))
                rels_changed = True
            if target:
                drawing_parts.add(_resolve_part(sheet_path, target))

        if rels_changed or rel_path not in out_entries:
            out_entries[rel_path] = ET.tostring(out_rels, encoding="utf-8", xml_declaration=True)

    for drawing_part in sorted(drawing_parts):
        if drawing_part not in template_entries:
            continue
        data = template_entries[drawing_part]
        if drawing_part.endswith(".xml"):
            data = _apply_date_to_drawing_xml(data, certificate_date)
        out_entries[drawing_part] = data

        rel_part = _worksheet_rel_path(drawing_part)
        if rel_part in template_entries:
            out_entries[rel_part] = template_entries[rel_part]

    tpl_ct = ET.fromstring(template_entries["[Content_Types].xml"])
    out_ct = ET.fromstring(out_entries["[Content_Types].xml"])
    existing_parts = {
        node.attrib.get("PartName", "")
        for node in out_ct.findall("{http://schemas.openxmlformats.org/package/2006/content-types}Override")
    }
    ct_changed = False
    for node in tpl_ct.findall("{http://schemas.openxmlformats.org/package/2006/content-types}Override"):
        part_name = node.attrib.get("PartName", "")
        if not part_name.startswith("/xl/drawings/"):
            continue
        if part_name in existing_parts:
            continue
        out_ct.append(ET.fromstring(ET.tostring(node, encoding="utf-8")))
        ct_changed = True
    if ct_changed:
        out_entries["[Content_Types].xml"] = ET.tostring(out_ct, encoding="utf-8", xml_declaration=True)

    temp_path = output_path.with_suffix(".tmp.xlsx")
    with ZipFile(temp_path, "w", ZIP_DEFLATED) as z_new:
        for name, data in out_entries.items():
            z_new.writestr(name, data)
    temp_path.replace(output_path)


def generate_certificates(
    template_path,
    csv_path,
    faculty,
    program,
    config_path="config.yml",
    progress_cb=None,
    certificate_date="",
):
    config = _load_config(config_path)
    course_metadata = _load_course_metadata(config, config_path)
    name_overrides, name_overrides_path = _load_name_overrides(config, config_path)
    df = read_csv(csv_path)

    missing_cols = validate_required_columns(df.columns)
    if missing_cols:
        raise ValueError("Faltan columnas requeridas: {0}".format(", ".join(missing_cols)))

    df = df.copy()
    df["DNI"] = df["DNI"].fillna("").astype(str).str.strip()
    df["CODIGO_CURSO"] = df["CODIGO"].apply(extract_course_code)
    df["ESTADO_N"] = df["ESTADO"].apply(normalize_str)
    df["TIPO_CURSO_N"] = df["TIPO_CURSO"].apply(normalize_str)
    df["NOMBRE_COMPLETO"] = df["NOMBRE_COMPLETO"].fillna("").astype(str).str.strip()
    df["CODIGO_ALUMNO"] = df["CODIGO_ALUMNO"].fillna("").astype(str).str.strip()
    resolution_flag_col = _resolve_column(
        df,
        None,
        ["RESOLUCION_FLAG", "RESOLUCION", "RESOLUCIÓN", "RESOLUCIO"],
    )
    if resolution_flag_col:
        df["_RESOLUTION_FLAG"] = df[resolution_flag_col]
    else:
        df["_RESOLUTION_FLAG"] = ""

    cara_cfg = config.get("cara", {})
    sello_cfg = config.get("sello", {})
    cara_sheet = str(cara_cfg.get("sheet_name", "CARA")).strip() or "CARA"
    sello_sheet = str(sello_cfg.get("sheet_name", "SELLO")).strip() or "SELLO"
    cara_cols = _get_course_columns(cara_cfg)
    sello_cols = _get_course_columns(sello_cfg)

    from openpyxl.utils.cell import column_index_from_string, get_column_letter

    def _col_index_required(col_letter: str, label: str) -> int:
        if not col_letter:
            raise ValueError("Columna faltante en configuración: {0}".format(label))
        return column_index_from_string(col_letter)

    def _col_index_optional(col_letter: str) -> int | None:
        if not col_letter:
            return None
        return column_index_from_string(col_letter)

    columns_by_sheet = {
        cara_sheet: {
            "code_col": _col_index_required(cara_cols["code_col"], "cara.course_columns.code_col"),
            "curso_col": _col_index_required(cara_cols["curso_col"], "cara.course_columns.curso_col"),
            "nota_col": _col_index_required(cara_cols["nota_col"], "cara.course_columns.nota_col"),
            "nota_texto_col": _col_index_required(
                cara_cols["nota_texto_col"], "cara.course_columns.nota_texto_col"
            ),
            "creditos_col": _col_index_optional(cara_cols["creditos_col"]),
            "acta_col": _col_index_optional(cara_cols["acta_col"]),
            "resolucion_col": _col_index_optional(cara_cols["resolucion_col"]),
            "x_col": _col_index_optional(cara_cols["x_col"]),
            "y_col": _col_index_optional(cara_cols["y_col"]),
            "z_col": _col_index_optional(cara_cols["z_col"]),
        },
        sello_sheet: {
            "code_col": _col_index_required(sello_cols["code_col"], "sello.course_columns.code_col"),
            "curso_col": _col_index_required(sello_cols["curso_col"], "sello.course_columns.curso_col"),
            "nota_col": _col_index_required(sello_cols["nota_col"], "sello.course_columns.nota_col"),
            "nota_texto_col": _col_index_required(
                sello_cols["nota_texto_col"], "sello.course_columns.nota_texto_col"
            ),
            "creditos_col": _col_index_optional(sello_cols["creditos_col"]),
            "acta_col": _col_index_optional(sello_cols["acta_col"]),
            "resolucion_col": _col_index_optional(sello_cols["resolucion_col"]),
            "x_col": _col_index_optional(sello_cols["x_col"]),
            "y_col": _col_index_optional(sello_cols["y_col"]),
            "z_col": _col_index_optional(sello_cols["z_col"]),
        },
    }

    def _validate_column_overlaps(sheet_name: str, cols: dict[str, int | None]) -> None:
        used: dict[int, str] = {}
        for key, col_idx in cols.items():
            if col_idx is None:
                continue
            prev = used.get(col_idx)
            if prev and prev != key:
                col_letter = get_column_letter(col_idx)
                raise ValueError(
                    "Columnas en conflicto en {0}: {1} y {2} usan {3}".format(
                        sheet_name, prev, key, col_letter
                    )
                )
            used[col_idx] = key

    _validate_column_overlaps(cara_sheet, columns_by_sheet[cara_sheet])
    _validate_column_overlaps(sello_sheet, columns_by_sheet[sello_sheet])

    slots_by_sheet = build_model_index(template_path, config)

    electives_cfg = config.get("electives", {})
    elective_years = set(electives_cfg.get("years", [6, 7, 8]))
    approval_years = set(electives_cfg.get("approval_only_years") or range(1, 9))

    with open(template_path, "rb") as f:
        template_bytes = f.read()

    out_dir = Path(tempfile.mkdtemp())
    cert_dir = out_dir / "certificados"
    cert_dir.mkdir(parents=True, exist_ok=True)

    logger = LogCollector()

    if df.empty:
        log_path = out_dir / "log.xlsx"
        log_df = logger.to_excel(log_path)
        zip_path = out_dir / "certificados.zip"
        zip_directory(out_dir, zip_path)
        return zip_path, log_df

    total_students = int(df["DNI"].nunique())
    processed_students = 0

    ambiguous_names: list[dict] = []

    for dni, grp in df.groupby("DNI"):
        errors = []
        dni_str = str(dni).strip()

        codigo_alumno = ""
        nombre = ""
        nombre_formateado = ""
        if not grp.empty:
            codigo_alumno = str(grp["CODIGO_ALUMNO"].iloc[0]).strip()
            nombre = str(grp["NOMBRE_COMPLETO"].iloc[0]).strip()
            # Validación de comas deshabilitada temporalmente (ver _format_person_name).
            # nombre_formateado, nombre_error = _format_person_name(
            #     nombre, dni_str, name_overrides, ambiguous_names
            # )
            # if nombre_error:
            #     errors.append(nombre_error)
            nombre_formateado, nombre_error = _format_simple_person_name(nombre)
            if nombre_error:
                errors.append(nombre_error)

        if not _is_valid_dni(dni_str):
            errors.append("DNI vacío o no numérico")

        student_resolution_raw = ""
        if "_RESOLUTION_FLAG" in grp.columns:
            for value in grp["_RESOLUTION_FLAG"].tolist():
                if str(value or "").strip():
                    student_resolution_raw = value
                    break
        student_use_resolution, student_resolution_error = _parse_resolution_flag(
            student_resolution_raw
        )
        if student_resolution_error:
            errors.append(student_resolution_error)

        required_fields = [
            "DNI",
            "NOMBRE_COMPLETO",
            "CODIGO",
            "CURSO",
            "YEAR",
            "TIPO_CURSO",
            "PERIODO",
            "ESTADO",
            "CODIGO_ALUMNO",
        ]

        records: list[dict] = []
        records_all: list[dict] = []
        for rec in grp.to_dict("records"):
            rec["CODIGO_CURSO"] = extract_course_code(rec.get("CODIGO"))
            rec["GRADE"] = parse_grade(rec.get("NOTA"))
            rec["YEAR_INT"] = parse_year(rec.get("YEAR"))
            rec["ESTADO_N"] = normalize_str(rec.get("ESTADO"))
            rec["USA_RESOLUCION"] = student_use_resolution
            rec["RESOLUCION_FLAG_ERROR"] = student_resolution_error
            rec["CURSO_NOMBRE"] = str(rec.get("CURSO") or "").strip()
            rec["CREDITOS"] = str(rec.get("CREDITOS") or "").strip()
            rec["ACTA"] = str(rec.get("ACTA") or "").strip()
            meta = course_metadata.get(rec["CODIGO_CURSO"], {})
            rec["RESOLUCION_TEXTO"] = str(meta.get("resolucion") or "").strip()
            rec["Z"] = str(meta.get("z") or "").strip()
            rec["PERIODO_VAL"] = _format_period_value(rec.get("PERIODO"))
            rec["CORRELATIVO"] = _resolve_correlativo(rec)

            records_all.append(rec)

            year_int = rec.get("YEAR_INT")
            if year_int in approval_years and rec["ESTADO_N"] != "APROBADO":
                continue

            records.append(rec)

        period_fallback_by_year: dict[int, str] = {}
        for rec in records:
            year_int = rec.get("YEAR_INT")
            if year_int not in (9, 10):
                continue
            period_val = rec.get("PERIODO_VAL") or ""
            if period_val and year_int not in period_fallback_by_year:
                period_fallback_by_year[year_int] = period_val

        for rec in records:
            year_int = rec.get("YEAR_INT")
            if year_int in (9, 10) and _is_missing(rec.get("PERIODO_VAL")):
                fallback = period_fallback_by_year.get(year_int, "")
                if fallback:
                    rec["PERIODO_VAL"] = fallback

            for field in required_fields:
                if _is_missing(rec.get(field)):
                    errors.append("Campo faltante: {0}".format(field))

            if _is_missing(rec.get("CODIGO_CURSO")):
                errors.append("Código de curso vacío")
            if rec.get("YEAR_INT") is None:
                errors.append("YEAR inválido")
            if _is_missing(rec.get("CURSO_NOMBRE")):
                errors.append("Curso vacío")
            if _is_missing(rec.get("PERIODO_VAL")):
                errors.append("PERIODO inválido")
            if rec.get("RESOLUCION_FLAG_ERROR"):
                errors.append(rec["RESOLUCION_FLAG_ERROR"])
            if _is_missing(rec.get("CORRELATIVO")):
                errors.append("Sin correlativo (Z) para curso: {0}".format(rec.get("CODIGO_CURSO")))
            if rec.get("USA_RESOLUCION", True) and _is_missing(rec.get("RESOLUCION_TEXTO")):
                errors.append("Sin resolución para curso: {0}".format(rec.get("CODIGO_CURSO")))

        def _deduplicate_records(items: list[dict]) -> list[dict]:
            best: dict[tuple[int | None, str], dict] = {}
            for rec in items:
                year = rec.get("YEAR_INT")
                code = rec.get("CODIGO_CURSO") or ""
                period = parse_period(rec.get("PERIODO"))
                grade = rec.get("GRADE")
                type_rank = 1 if normalize_str(rec.get("TIPO_CURSO")) == "ELECTIVO" else 0
                rank = (
                    type_rank,
                    period if period is not None else -1,
                    grade if grade is not None else -1,
                )
                key = (year, code)
                current = best.get(key)
                if current is None or rank > current["_rank"]:
                    rec["_rank"] = rank
                    best[key] = rec
            deduped = list(best.values())
            for rec in deduped:
                rec.pop("_rank", None)
            return deduped

        if errors:
            reason = "; ".join(sorted(set(errors)))
            logger.add(dni_str, codigo_alumno, nombre, "ERROR", reason, "")
            processed_students += 1
            if progress_cb:
                progress_cb(processed_students, total_students)
            continue

        records = _deduplicate_records(records)

        approved_grade_by_code: dict[str, int] = {}
        for rec in records:
            code = str(rec.get("CODIGO_CURSO") or "").strip()
            grade = rec.get("GRADE")
            if not code:
                continue
            if grade is None or grade == 0:
                continue
            approved_grade_by_code[code] = int(grade)

        electives = [
            rec
            for rec in records
            if normalize_str(rec.get("TIPO_CURSO")) == "ELECTIVO"
            and rec.get("YEAR_INT") in elective_years
        ]
        if len(electives) != 3:
            logger.add(
                dni_str,
                codigo_alumno,
                nombre,
                "ERROR",
                "El alumno tiene {0} electivos aprobados (se requieren 3)".format(len(electives)),
                "",
            )
            processed_students += 1
            if progress_cb:
                progress_cb(processed_students, total_students)
            continue

        balanced, balance_error = balance_electives(electives, years=elective_years)
        if balance_error:
            logger.add(dni_str, codigo_alumno, nombre, "ERROR", balance_error, "")
            processed_students += 1
            if progress_cb:
                progress_cb(processed_students, total_students)
            continue

        for rec in balanced:
            rec["YEAR_INT"] = rec.get("TARGET_YEAR")

        electives_by_year: dict[int, list[dict]] = {}
        fixed_by_year: dict[int, list[dict]] = {}
        years_with_records: set[int] = set()

        for rec in records:
            year = rec.get("YEAR_INT")
            if year is None:
                continue
            years_with_records.add(year)
            if normalize_str(rec.get("TIPO_CURSO")) == "ELECTIVO" and year in elective_years:
                electives_by_year.setdefault(year, []).append(rec)
            else:
                fixed_by_year.setdefault(year, []).append(rec)

        missing_reasons = []
        for year in sorted(years_with_records):
            target_sheet = _select_sheet(year, cara_sheet, sello_sheet)
            if not target_sheet:
                missing_reasons.append("YEAR fuera de rango (1-10): {0}".format(year))
                continue
            sheet_slots = slots_by_sheet.get(target_sheet, {})
            rows = sheet_slots.get(year)
            if not rows:
                missing_reasons.append(
                    "No hay filas para YEAR={0} en {1}".format(year, target_sheet)
                )
                continue
            electives_for_year = electives_by_year.get(year, [])
            if electives_for_year and len(rows) < len(electives_for_year):
                missing_reasons.append(
                    "No hay suficientes filas en {0} para YEAR={1}: {2} electivos, {3} filas".format(
                        target_sheet, year, len(electives_for_year), len(rows)
                    )
                )

        if missing_reasons:
            reason = "; ".join(sorted(set(missing_reasons)))
            logger.add(dni_str, codigo_alumno, nombre, "ERROR", reason, "")
            processed_students += 1
            if progress_cb:
                progress_cb(processed_students, total_students)
            continue

        wb = load_workbook(BytesIO(template_bytes))
        identity = {
            "DNI": dni_str,
            "NOMBRE": nombre_formateado or nombre,
            "FACULTAD": faculty,
            "PROGRAMA": program,
            "FECHA": certificate_date,
        }

        try:
            if cara_sheet not in wb.sheetnames:
                raise ValueError("La hoja '{0}' no existe en la plantilla".format(cara_sheet))
            ws = wb[cara_sheet]
            _write_identity(ws, identity, cara_cfg.get("header_cells", {}))
        except Exception as exc:
            logger.add(dni_str, codigo_alumno, nombre, "ERROR", str(exc), "")
            processed_students += 1
            if progress_cb:
                progress_cb(processed_students, total_students)
            continue

        def _elective_code_key(rec: dict) -> str:
            return str(rec.get("CODIGO_CURSO") or rec.get("CODIGO") or "").strip().upper()

        name_by_code: dict[str, str] = {}
        for rec in records_all:
            code = str(rec.get("CODIGO_CURSO") or "").strip()
            name = _normalize_spaces(rec.get("CURSO_NOMBRE") or "")
            if code and name and code not in name_by_code:
                name_by_code[code] = name

        template_codes_by_year: dict[int, set[str]] = {}
        for sheet_name, sheet_slots in slots_by_sheet.items():
            ws = wb[sheet_name]
            code_col = columns_by_sheet[sheet_name]["code_col"]
            for year, rows in sheet_slots.items():
                for row in rows:
                    code = extract_course_code(ws.cell(row=row, column=code_col).value)
                    if code:
                        template_codes_by_year.setdefault(year, set()).add(code)

        electives_used_by_year: dict[int, list[dict]] = {}
        for year in elective_years:
            electives_used_by_year[year] = sorted(
                electives_by_year.get(year, []), key=_elective_code_key
            )

        observed_codes: set[str] = set()
        for year, codes in template_codes_by_year.items():
            for code in codes:
                if not code:
                    continue
                if code in approved_grade_by_code:
                    continue
                has_missing = False
                for rec in records_all:
                    if str(rec.get("CODIGO_CURSO") or "").strip() != code:
                        continue
                    grade = rec.get("GRADE")
                    if grade is None or grade == 0:
                        has_missing = True
                        break
                if has_missing:
                    observed_codes.add(code)

        elective_codes: set[str] = set()
        for rec in records_all:
            if normalize_str(rec.get("TIPO_CURSO")) != "ELECTIVO":
                continue
            code = str(rec.get("CODIGO_CURSO") or "").strip()
            if code:
                elective_codes.add(code)
        observed_codes.difference_update(elective_codes)

        observed_courses: list[str] = []
        for code in sorted(observed_codes):
            name = name_by_code.get(code, "")
            if name:
                observed_courses.append("{0} - {1}".format(code, name))
            else:
                observed_courses.append(code)

        for target_sheet, sheet_slots in slots_by_sheet.items():
            ws = wb[target_sheet]
            merged_ranges = list(ws.merged_cells.ranges)
            cols = columns_by_sheet[target_sheet]

            for year, rows in sheet_slots.items():
                if not rows:
                    continue

                fixed_records = fixed_by_year.get(year, [])
                fixed_by_code = {rec.get("CODIGO_CURSO"): rec for rec in fixed_records}

                electives_sorted = electives_used_by_year.get(year, [])
                elective_count = len(electives_sorted)
                elective_rows = rows[-elective_count:] if elective_count else []
                fixed_rows = rows[:-elective_count] if elective_count else rows

                for row in fixed_rows:
                    template_code = extract_course_code(
                        ws.cell(row=row, column=cols["code_col"]).value
                    )
                    rec = fixed_by_code.get(template_code)
                    grade = _grade_for_output(rec.get("GRADE") if rec else None)
                    _safe_set_cell(ws, row, cols["nota_col"], grade, merged_ranges)
                    if rec:
                        x_col = cols.get("x_col")
                        y_col = cols.get("y_col")
                        z_col = cols.get("z_col")
                        x_ref = "{0}{1}".format(get_column_letter(x_col), row) if x_col else None
                        y_ref = "{0}{1}".format(get_column_letter(y_col), row) if y_col else None
                        z_ref = "{0}{1}".format(get_column_letter(z_col), row) if z_col else None
                        _safe_set_cell(
                            ws,
                            row,
                            cols.get("resolucion_col"),
                            _resolution_output_value(rec, x_ref, y_ref, z_ref),
                            merged_ranges,
                        )
                        _safe_set_cell(
                            ws, row, cols.get("x_col"), rec.get("PERIODO_VAL", ""), merged_ranges
                        )
                        _safe_set_cell(ws, row, cols.get("y_col"), "21", merged_ranges)
                        _safe_set_cell(
                            ws, row, cols.get("z_col"), rec.get("CORRELATIVO", ""), merged_ranges
                        )

                for row, rec in zip(elective_rows, electives_sorted):
                    grade = _grade_for_output(rec.get("GRADE"))
                    x_col = cols.get("x_col")
                    y_col = cols.get("y_col")
                    z_col = cols.get("z_col")
                    x_ref = "{0}{1}".format(get_column_letter(x_col), row) if x_col else None
                    y_ref = "{0}{1}".format(get_column_letter(y_col), row) if y_col else None
                    z_ref = "{0}{1}".format(get_column_letter(z_col), row) if z_col else None
                    _safe_set_cell(ws, row, cols["code_col"], rec.get("CODIGO_CURSO"), merged_ranges)
                    _safe_set_cell(
                        ws,
                        row,
                        cols["curso_col"],
                        _format_course_name(rec.get("CURSO_NOMBRE", "")),
                        merged_ranges,
                    )
                    _safe_set_cell(ws, row, cols["nota_col"], grade, merged_ranges)
                    _safe_set_cell(
                        ws, row, cols.get("creditos_col"), rec.get("CREDITOS", ""), merged_ranges
                    )
                    _safe_set_cell(ws, row, cols.get("acta_col"), rec.get("ACTA", ""), merged_ranges)
                    _safe_set_cell(
                        ws,
                        row,
                        cols.get("resolucion_col"),
                        _resolution_output_value(rec, x_ref, y_ref, z_ref),
                        merged_ranges,
                    )
                    _safe_set_cell(
                        ws, row, cols.get("x_col"), rec.get("PERIODO_VAL", ""), merged_ranges
                    )
                    _safe_set_cell(ws, row, cols.get("y_col"), "21", merged_ranges)
                    _safe_set_cell(
                        ws, row, cols.get("z_col"), rec.get("CORRELATIVO", ""), merged_ranges
                    )

        safe_name = _sanitize_filename("{0} {1} - {2}".format(dni_str, codigo_alumno, nombre))
        if not safe_name:
            safe_name = dni_str or "alumno"

        out_path = cert_dir / "{0}.xlsx".format(safe_name)
        wb.save(out_path)
        _copy_template_drawings(Path(template_path), out_path, certificate_date)

        status = "OK"
        reason = ""
        if observed_courses:
            status = "OBSERVADO"
            reason = "Cursos con nota 0 o sin nota: {0}".format("; ".join(observed_courses))
        logger.add(dni_str, codigo_alumno, nombre, status, reason, out_path.name)
        processed_students += 1
        if progress_cb:
            progress_cb(processed_students, total_students)

    log_path = out_dir / "log.xlsx"
    log_df = logger.to_excel(log_path)

    zip_path = out_dir / "certificados.zip"
    zip_directory(out_dir, zip_path)

    if ambiguous_names and name_overrides_path:
        _write_name_overrides(name_overrides_path, ambiguous_names)

    return zip_path, log_df
