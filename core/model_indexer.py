import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

ORDINAL_MAP = {
    "PRIMER": 1,
    "PRIMERO": 1,
    "SEGUNDO": 2,
    "TERCER": 3,
    "TERCERO": 3,
    "CUARTO": 4,
    "QUINTO": 5,
    "SEXTO": 6,
    "SEPTIMO": 7,
    "OCTAVO": 8,
    "NOVENO": 9,
    "DECIMO": 10,
}


def _normalize_text(text: object) -> str:
    if text is None:
        return ""
    value = str(text).strip().upper()
    value = unicodedata.normalize("NFD", value)
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    value = re.sub(r"[^A-Z0-9]+", "", value)
    return value


def _extract_cycle_number(text: object) -> int | None:
    norm = _normalize_text(text)
    if not norm or "CICLO" not in norm:
        return None
    digits = re.findall(r"\d+", norm)
    if digits:
        return int(digits[0])
    for key, num in ORDINAL_MAP.items():
        if key in norm:
            return num
    return None


def _scan_cycle_headers(ws, code_col_letter: str) -> list[tuple[int, int]]:
    col_idx = column_index_from_string(code_col_letter)
    headers: list[tuple[int, int]] = []
    seen_years = set()
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        cell = row[0]
        year = _extract_cycle_number(cell.value)
        if year is None:
            continue
        if year in seen_years:
            continue
        headers.append((year, cell.row))
        seen_years.add(year)
    headers.sort(key=lambda item: item[1])
    return headers


def _build_slots(headers: list[tuple[int, int]], max_row: int) -> dict[int, list[int]]:
    slots: dict[int, list[int]] = {}
    for idx, (year, row) in enumerate(headers):
        start = row + 1
        end = max_row if idx + 1 >= len(headers) else headers[idx + 1][1] - 1
        if end < start:
            slots[year] = []
            continue
        slots[year] = list(range(start, end + 1))
    return slots


def _get_sheet_name(cfg: dict, fallback: str) -> str:
    name = cfg.get("sheet_name") or cfg.get("name") or fallback
    return str(name).strip()


def _get_code_col(cfg: dict) -> str:
    cols = cfg.get("course_columns") or {}
    return str(cols.get("code_col", "B")).strip().upper()


def build_model_index(template_path, config):
    wb = load_workbook(template_path, data_only=True)

    cara_cfg = config.get("cara", {})
    sello_cfg = config.get("sello", {})

    slots_by_sheet: dict[str, dict[int, list[int]]] = {}
    for cfg, fallback in ((cara_cfg, "CARA"), (sello_cfg, "SELLO")):
        sheet_name = _get_sheet_name(cfg, fallback)
        if sheet_name not in wb.sheetnames:
            raise ValueError("La hoja '{0}' no existe en la plantilla".format(sheet_name))
        ws = wb[sheet_name]
        code_col = _get_code_col(cfg)
        headers = _scan_cycle_headers(ws, code_col)
        if not headers:
            raise ValueError("No se encontraron filas de CICLO en la hoja '{0}'".format(sheet_name))
        slots_by_sheet[sheet_name] = _build_slots(headers, ws.max_row)

    return slots_by_sheet
